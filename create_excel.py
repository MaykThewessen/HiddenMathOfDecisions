"""Generate the Decision Math Excel workbook with all 6 formula examples + dashboard."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True, size=14, color="2F5496")
LABEL_FONT = Font(bold=True, size=11)
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RESULT_FONT = Font(bold=True, size=11, color="375623")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
WARN_FONT = Font(bold=True, size=11, color="C00000")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN_BORDER


def style_label(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = LABEL_FONT
    cell.border = THIN_BORDER


def style_input(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt


def style_result(ws, row, col, value=None, formula=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=formula if formula else value)
    cell.fill = RESULT_FILL
    cell.font = RESULT_FONT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt


def style_warn(ws, row, col, value=None, formula=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=formula if formula else value)
    cell.fill = WARN_FILL
    cell.font = WARN_FONT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt


# ============================================================
# Sheet 1: Expected Value
# ============================================================
ws = wb.active
ws.title = "1 - Expected Value"
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18

ws.cell(row=1, column=1, value="1. Expected Value").font = SECTION_FONT
ws.cell(row=2, column=1, value="EV = Sum(probability x payoff)").font = Font(italic=True, size=11, color="666666")

# Coin flip example
ws.cell(row=4, column=1, value="Example: Coin Flip Bet").font = Font(bold=True, size=11)
for c, h in enumerate(["Outcome", "Probability", "Payoff ($)", "Weighted Value"], 1):
    style_header(ws, 5, c, h)

style_label(ws, 6, 1, "Heads (win)")
style_input(ws, 6, 2, 0.50, "0%")
style_input(ws, 6, 3, 150, "$#,##0")
ws.cell(row=6, column=4, value="=B6*C6").number_format = "$#,##0.00"
ws.cell(row=6, column=4).border = THIN_BORDER

style_label(ws, 7, 1, "Tails (lose)")
style_input(ws, 7, 2, 0.50, "0%")
style_input(ws, 7, 3, -100, "$#,##0")
ws.cell(row=7, column=4, value="=B7*C7").number_format = "$#,##0.00"
ws.cell(row=7, column=4).border = THIN_BORDER

style_label(ws, 8, 1, "EXPECTED VALUE")
style_result(ws, 8, 4, formula="=SUM(D6:D7)", fmt="$#,##0.00")

# Over 100 flips
style_label(ws, 9, 1, "Over 100 flips")
style_result(ws, 9, 4, formula="=D8*100", fmt="$#,##0")

# Career example
ws.cell(row=11, column=1, value="Example: Career Choice").font = Font(bold=True, size=11)
for c, h in enumerate(["Outcome", "Probability", "Payoff ($)", "Weighted Value"], 1):
    style_header(ws, 12, c, h)

style_label(ws, 13, 1, "Option A: Safe job")
style_input(ws, 13, 2, 1.00, "0%")
style_input(ws, 13, 3, 120000, "$#,##0")
ws.cell(row=13, column=4, value="=B13*C13").number_format = "$#,##0"
ws.cell(row=13, column=4).border = THIN_BORDER

style_label(ws, 14, 1, "EV of Option A")
style_result(ws, 14, 4, formula="=D13", fmt="$#,##0")

ws.cell(row=16, column=1, value="Option B: Startup").font = Font(bold=True)
style_label(ws, 17, 1, "Startup succeeds")
style_input(ws, 17, 2, 0.60, "0%")
style_input(ws, 17, 3, 250000, "$#,##0")
ws.cell(row=17, column=4, value="=B17*C17").number_format = "$#,##0"
ws.cell(row=17, column=4).border = THIN_BORDER

style_label(ws, 18, 1, "Startup fails")
style_input(ws, 18, 2, 0.40, "0%")
style_input(ws, 18, 3, 70000, "$#,##0")
ws.cell(row=18, column=4, value="=B18*C18").number_format = "$#,##0"
ws.cell(row=18, column=4).border = THIN_BORDER

style_label(ws, 19, 1, "EV of Option B")
style_result(ws, 19, 4, formula="=SUM(D17:D18)", fmt="$#,##0")

style_label(ws, 21, 1, "B is better by:")
style_result(ws, 21, 4, formula="=D19-D14", fmt="$#,##0")

# Chart: Compare EV of A vs B
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Expected Value: Option A vs B"
chart1.y_axis.title = "Expected Value ($)"
chart1.y_axis.numFmt = "$#,##0"

# We need helper data for the chart
ws.cell(row=23, column=1, value="Option A").font = Font(size=9, color="999999")
ws.cell(row=24, column=1, value="Option B").font = Font(size=9, color="999999")
ws.cell(row=23, column=2, value="=D14").number_format = "$#,##0"
ws.cell(row=24, column=2, value="=D19").number_format = "$#,##0"

cats = Reference(ws, min_col=1, min_row=23, max_row=24)
vals = Reference(ws, min_col=2, min_row=23, max_row=24)
chart1.add_data(vals, titles_from_data=False)
chart1.set_categories(cats)
chart1.series[0].title = openpyxl.chart.series.SeriesLabel(v="Expected Value")
chart1.series[0].graphicalProperties.solidFill = "2F5496"
chart1.width = 15
chart1.height = 10
ws.add_chart(chart1, "F4")

# ============================================================
# Sheet 2: Base Rate Neglect
# ============================================================
ws2 = wb.create_sheet("2 - Base Rate Neglect")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 18

ws2.cell(row=1, column=1, value="2. Base Rate Neglect").font = SECTION_FONT
ws2.cell(row=2, column=1, value="P(A|B) = P(B|A) x P(A) / P(B)").font = Font(italic=True, size=11, color="666666")

ws2.cell(row=4, column=1, value="Disease Test Example").font = Font(bold=True, size=11)

style_label(ws2, 6, 1, "Base rate (prevalence)")
style_input(ws2, 6, 2, 0.001, "0.000%")

style_label(ws2, 7, 1, "Test sensitivity (true pos rate)")
style_input(ws2, 7, 2, 0.99, "0%")

style_label(ws2, 8, 1, "False positive rate")
style_input(ws2, 8, 2, 0.01, "0%")

style_label(ws2, 10, 1, "Per 1,000 people:")
style_label(ws2, 11, 1, "  True positives")
ws2.cell(row=11, column=2, value="=B6*1000*B7").number_format = "0.00"
ws2.cell(row=11, column=2).border = THIN_BORDER

style_label(ws2, 12, 1, "  False positives")
ws2.cell(row=12, column=2, value="=(1-B6)*1000*B8").number_format = "0.00"
ws2.cell(row=12, column=2).border = THIN_BORDER

style_label(ws2, 13, 1, "  Total positives")
ws2.cell(row=13, column=2, value="=B11+B12").number_format = "0.00"
ws2.cell(row=13, column=2).border = THIN_BORDER

style_label(ws2, 15, 1, "P(actually sick | positive test)")
style_result(ws2, 15, 2, formula="=B11/B13", fmt="0.0%")

style_label(ws2, 16, 1, "What most people guess")
style_warn(ws2, 16, 2, value=0.99, fmt="0.0%")

style_label(ws2, 17, 1, "How wrong they are")
style_warn(ws2, 17, 2, formula="=B16-B15", fmt="0.0%")

ws2.cell(row=19, column=1, value="The base rate dominates the result!").font = Font(
    italic=True, color="C00000"
)

# Chart: True vs False positives
chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.title = "Positive Test Results Breakdown"
chart2.y_axis.title = "Count per 1,000"

ws2.cell(row=21, column=1, value="True Positives").font = Font(size=9, color="999999")
ws2.cell(row=22, column=1, value="False Positives").font = Font(size=9, color="999999")
ws2.cell(row=21, column=2, value="=B11")
ws2.cell(row=22, column=2, value="=B12")

cats2 = Reference(ws2, min_col=1, min_row=21, max_row=22)
vals2 = Reference(ws2, min_col=2, min_row=21, max_row=22)
chart2.add_data(vals2, titles_from_data=False)
chart2.set_categories(cats2)
chart2.series[0].title = openpyxl.chart.series.SeriesLabel(v="Count")
# Color first bar green, second red
dp_green = DataPoint(idx=0)
dp_green.graphicalProperties.solidFill = "375623"
dp_red = DataPoint(idx=1)
dp_red.graphicalProperties.solidFill = "C00000"
chart2.series[0].data_points = [dp_green, dp_red]
chart2.width = 15
chart2.height = 10
ws2.add_chart(chart2, "D4")

# ============================================================
# Sheet 3: Sunk Cost
# ============================================================
ws3 = wb.create_sheet("3 - Sunk Cost Fallacy")
ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 18

ws3.cell(row=1, column=1, value="3. Sunk Cost Fallacy").font = SECTION_FONT
ws3.cell(row=2, column=1, value="Only compare FUTURE costs and benefits").font = Font(italic=True, size=11, color="666666")

ws3.cell(row=4, column=1, value="The Rational Decision Framework").font = Font(bold=True, size=11)

style_label(ws3, 6, 1, "Amount already invested (SUNK)")
style_input(ws3, 6, 2, 15, "$#,##0")

style_label(ws3, 7, 1, "Future cost to continue")
style_input(ws3, 7, 2, 0, "$#,##0")

style_label(ws3, 8, 1, "Future value if you CONTINUE")
style_input(ws3, 8, 2, -5, "$#,##0")

style_label(ws3, 9, 1, "Future value if you QUIT")
style_input(ws3, 9, 2, 10, "$#,##0")

style_label(ws3, 11, 1, "Net future value: CONTINUE")
style_result(ws3, 11, 2, formula="=B8-B7", fmt="$#,##0")

style_label(ws3, 12, 1, "Net future value: QUIT")
style_result(ws3, 12, 2, formula="=B9", fmt="$#,##0")

style_label(ws3, 14, 1, "RATIONAL DECISION")
style_result(ws3, 14, 2, formula='=IF(B11>B12,"CONTINUE","QUIT")')

ws3.cell(row=16, column=1, value="Note: The sunk cost is IRRELEVANT").font = Font(italic=True, color="C00000")
ws3.cell(row=17, column=1, value="to the forward-looking decision.").font = Font(italic=True, color="C00000")

# Sunk cost visual
ws3.cell(row=19, column=1, value="Try changing the sunk cost above.").font = Font(bold=True, size=11)
ws3.cell(row=20, column=1, value="Notice the decision NEVER changes!").font = Font(bold=True, size=11, color="2F5496")
ws3.cell(row=21, column=1, value="That proves it's irrelevant.").font = Font(italic=True, color="666666")

# ============================================================
# Sheet 4: Bayesian Thinking
# ============================================================
ws4 = wb.create_sheet("4 - Bayesian Thinking")
ws4.column_dimensions["A"].width = 35
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 18
ws4.column_dimensions["D"].width = 18

ws4.cell(row=1, column=1, value="4. Bayesian Thinking").font = SECTION_FONT
ws4.cell(row=2, column=1, value="P(H|E) = P(E|H) x P(H) / P(E)").font = Font(italic=True, size=11, color="666666")

ws4.cell(row=4, column=1, value="Is your coworker about to quit?").font = Font(bold=True, size=11)

for c, h in enumerate(["Stage", "P(evidence|quit)", "P(evidence|stay)", "Updated P(quit)"], 1):
    style_header(ws4, 6, c, h)

style_label(ws4, 7, 1, "Prior (before evidence)")
ws4.cell(row=7, column=4, value=0.10).number_format = "0.0%"
ws4.cell(row=7, column=4).fill = INPUT_FILL
ws4.cell(row=7, column=4).border = THIN_BORDER

style_label(ws4, 8, 1, "Evidence 1: LinkedIn update")
style_input(ws4, 8, 2, 0.70, "0%")
style_input(ws4, 8, 3, 0.15, "0%")
style_result(ws4, 8, 4, formula="=(B8*D7)/(B8*D7+C8*(1-D7))", fmt="0.0%")

style_label(ws4, 9, 1, "Evidence 2: Outside phone calls")
style_input(ws4, 9, 2, 0.60, "0%")
style_input(ws4, 9, 3, 0.10, "0%")
style_result(ws4, 9, 4, formula="=(B9*D8)/(B9*D8+C9*(1-D8))", fmt="0.0%")

style_label(ws4, 10, 1, "Evidence 3: Vague on projects")
style_input(ws4, 10, 2, 0.80, "0%")
style_input(ws4, 10, 3, 0.20, "0%")
style_result(ws4, 10, 4, formula="=(B10*D9)/(B10*D9+C10*(1-D9))", fmt="0.0%")

# Add rows for more evidence
style_label(ws4, 11, 1, "Evidence 4: (add your own)")
style_input(ws4, 11, 2, 0.50, "0%")
style_input(ws4, 11, 3, 0.50, "0%")
style_result(ws4, 11, 4, formula="=(B11*D10)/(B11*D10+C11*(1-D10))", fmt="0.0%")

style_label(ws4, 12, 1, "Evidence 5: (add your own)")
style_input(ws4, 12, 2, 0.50, "0%")
style_input(ws4, 12, 3, 0.50, "0%")
style_result(ws4, 12, 4, formula="=(B12*D11)/(B12*D11+C12*(1-D11))", fmt="0.0%")

ws4.cell(row=14, column=1, value="Each piece of evidence nudges — never jumps.").font = Font(
    italic=True, color="2F5496"
)

# Chart: Bayesian update progression
chart4 = LineChart()
chart4.title = "Belief Update Progression"
chart4.y_axis.title = "P(quitting)"
chart4.y_axis.numFmt = "0%"
chart4.x_axis.title = "Evidence Stage"
chart4.style = 10

# Helper data for chart
for i, label in enumerate(["Prior", "LinkedIn", "Calls", "Vague", "Ev4", "Ev5"], start=0):
    ws4.cell(row=16 + i, column=1, value=label).font = Font(size=9, color="999999")
    ws4.cell(row=16 + i, column=2, value=f"=D{7 + i}").number_format = "0.0%"

cats4 = Reference(ws4, min_col=1, min_row=16, max_row=21)
vals4 = Reference(ws4, min_col=2, min_row=16, max_row=21)
chart4.add_data(vals4, titles_from_data=False)
chart4.set_categories(cats4)
chart4.series[0].title = openpyxl.chart.series.SeriesLabel(v="P(quit)")
chart4.series[0].graphicalProperties.line.solidFill = "2F5496"
chart4.width = 18
chart4.height = 10
ws4.add_chart(chart4, "F4")

# ============================================================
# Sheet 5: Survivorship Bias
# ============================================================
ws5 = wb.create_sheet("5 - Survivorship Bias")
ws5.column_dimensions["A"].width = 35
ws5.column_dimensions["B"].width = 18

ws5.cell(row=1, column=1, value="5. Survivorship Bias").font = SECTION_FONT
ws5.cell(row=2, column=1, value="Real rate = successes / total attempts").font = Font(italic=True, size=11, color="666666")

ws5.cell(row=4, column=1, value="Crypto Twitter Example").font = Font(bold=True, size=11)

style_label(ws5, 6, 1, "Visible success stories")
style_input(ws5, 6, 2, 50)

style_label(ws5, 7, 1, "Total attempts (estimated)")
style_input(ws5, 7, 2, 500)

style_label(ws5, 9, 1, "Invisible failures")
style_result(ws5, 9, 2, formula="=B7-B6")

style_label(ws5, 10, 1, "REAL success rate")
style_result(ws5, 10, 2, formula="=B6/B7", fmt="0.0%")

style_label(ws5, 11, 1, "What your feed shows")
style_warn(ws5, 11, 2, formula="=B6/(B6)", fmt="0.0%")

ws5.cell(row=13, column=1, value="Spotify Example").font = Font(bold=True, size=11)
style_label(ws5, 14, 1, "Daily uploads")
style_input(ws5, 14, 2, 90000)
style_label(ws5, 15, 1, "Tracks that 'go viral'")
style_input(ws5, 15, 2, 10)
style_label(ws5, 16, 1, "Real viral rate")
style_result(ws5, 16, 2, formula="=B15/B14", fmt="0.0000%")

ws5.cell(row=18, column=1, value="Restaurant Example").font = Font(bold=True, size=11)
style_label(ws5, 19, 1, "Restaurants opened")
style_input(ws5, 19, 2, 1000)
style_label(ws5, 20, 1, "Survive 1 year")
style_result(ws5, 20, 2, formula="=B19*0.4", fmt="#,##0")
style_label(ws5, 21, 1, "Survive 5 years")
style_result(ws5, 21, 2, formula="=B19*0.2", fmt="#,##0")
style_label(ws5, 22, 1, "5-year survival rate")
style_result(ws5, 22, 2, formula="=B21/B19", fmt="0.0%")

ws5.cell(row=24, column=1, value="Always look for the DENOMINATOR.").font = Font(
    italic=True, color="C00000"
)

# Chart: visible vs invisible
chart5 = BarChart()
chart5.type = "col"
chart5.style = 10
chart5.title = "What You See vs Reality (Crypto)"

ws5.cell(row=26, column=1, value="Visible (winners)").font = Font(size=9, color="999999")
ws5.cell(row=27, column=1, value="Invisible (losers)").font = Font(size=9, color="999999")
ws5.cell(row=26, column=2, value="=B6")
ws5.cell(row=27, column=2, value="=B9")

cats5 = Reference(ws5, min_col=1, min_row=26, max_row=27)
vals5 = Reference(ws5, min_col=2, min_row=26, max_row=27)
chart5.add_data(vals5, titles_from_data=False)
chart5.set_categories(cats5)
chart5.series[0].title = openpyxl.chart.series.SeriesLabel(v="People")
dp5_green = DataPoint(idx=0)
dp5_green.graphicalProperties.solidFill = "375623"
dp5_red = DataPoint(idx=1)
dp5_red.graphicalProperties.solidFill = "C00000"
chart5.series[0].data_points = [dp5_green, dp5_red]
chart5.width = 15
chart5.height = 10
ws5.add_chart(chart5, "D4")

# ============================================================
# Sheet 6: Kelly Criterion
# ============================================================
ws6 = wb.create_sheet("6 - Kelly Criterion")
ws6.column_dimensions["A"].width = 30
ws6.column_dimensions["B"].width = 18
ws6.column_dimensions["C"].width = 18

ws6.cell(row=1, column=1, value="6. Kelly Criterion").font = SECTION_FONT
ws6.cell(row=2, column=1, value="f* = (p x b - q) / b").font = Font(italic=True, size=11, color="666666")

ws6.cell(row=4, column=1, value="Inputs").font = Font(bold=True, size=11)

style_label(ws6, 5, 1, "Win probability (p)")
style_input(ws6, 5, 2, 0.60, "0%")

style_label(ws6, 6, 1, "Loss probability (q = 1-p)")
ws6.cell(row=6, column=2, value="=1-B5").number_format = "0%"
ws6.cell(row=6, column=2).border = THIN_BORDER

style_label(ws6, 7, 1, "Win multiplier (b)")
style_input(ws6, 7, 2, 1.0, "0.0")

style_label(ws6, 8, 1, "Bankroll ($)")
style_input(ws6, 8, 2, 10000, "$#,##0")

ws6.cell(row=10, column=1, value="Results").font = Font(bold=True, size=11)

style_label(ws6, 11, 1, "Edge (p x b - q)")
style_result(ws6, 11, 2, formula="=B5*B7-B6", fmt="0.0%")

style_label(ws6, 12, 1, "Full Kelly (f*)")
style_result(ws6, 12, 2, formula="=MAX(0,(B5*B7-B6)/B7)", fmt="0.0%")

style_label(ws6, 13, 1, "Half Kelly")
style_result(ws6, 13, 2, formula="=B12*0.5", fmt="0.0%")

style_label(ws6, 14, 1, "Quarter Kelly (recommended)")
style_result(ws6, 14, 2, formula="=B12*0.25", fmt="0.0%")

ws6.cell(row=16, column=1, value="Bet Sizes ($)").font = Font(bold=True, size=11)

for i, (label, ref) in enumerate([
    ("Full Kelly", "=B12*B8"),
    ("Half Kelly", "=B13*B8"),
    ("Quarter Kelly", "=B14*B8"),
], start=17):
    style_label(ws6, i, 1, label)
    style_result(ws6, i, 2, formula=ref, fmt="$#,##0")

ws6.cell(row=21, column=1, value="Full Kelly is too aggressive for humans.").font = Font(
    italic=True, color="C00000"
)
ws6.cell(row=22, column=1, value="Use quarter to half Kelly in practice.").font = Font(
    italic=True, color="C00000"
)

# Chart: Kelly fractions
chart6 = BarChart()
chart6.type = "col"
chart6.style = 10
chart6.title = "Bet Size by Kelly Fraction"
chart6.y_axis.title = "Bet Size ($)"
chart6.y_axis.numFmt = "$#,##0"

ws6.cell(row=24, column=1, value="Full Kelly").font = Font(size=9, color="999999")
ws6.cell(row=25, column=1, value="Half Kelly").font = Font(size=9, color="999999")
ws6.cell(row=26, column=1, value="Quarter Kelly").font = Font(size=9, color="999999")
ws6.cell(row=24, column=2, value="=B17")
ws6.cell(row=25, column=2, value="=B18")
ws6.cell(row=26, column=2, value="=B19")

cats6 = Reference(ws6, min_col=1, min_row=24, max_row=26)
vals6 = Reference(ws6, min_col=2, min_row=24, max_row=26)
chart6.add_data(vals6, titles_from_data=False)
chart6.set_categories(cats6)
chart6.series[0].title = openpyxl.chart.series.SeriesLabel(v="Bet Size")

dp6_red = DataPoint(idx=0)
dp6_red.graphicalProperties.solidFill = "C00000"
dp6_yellow = DataPoint(idx=1)
dp6_yellow.graphicalProperties.solidFill = "ED7D31"
dp6_green = DataPoint(idx=2)
dp6_green.graphicalProperties.solidFill = "375623"
chart6.series[0].data_points = [dp6_red, dp6_yellow, dp6_green]
chart6.width = 15
chart6.height = 10
ws6.add_chart(chart6, "D4")

# ============================================================
# Sheet 7: Dashboard — Your Decision Calculator
# ============================================================
ws7 = wb.create_sheet("Dashboard")
ws7.column_dimensions["A"].width = 35
ws7.column_dimensions["B"].width = 20
ws7.column_dimensions["C"].width = 20
ws7.column_dimensions["D"].width = 20
ws7.column_dimensions["E"].width = 20

ws7.cell(row=1, column=1, value="DECISION DASHBOARD").font = Font(bold=True, size=16, color="2F5496")
ws7.cell(row=2, column=1, value="Enter your scenario below. All 6 models run automatically.").font = Font(
    italic=True, size=11, color="666666"
)

# --- Section: Your Decision ---
ws7.cell(row=4, column=1, value="YOUR DECISION").font = Font(bold=True, size=13, color="2F5496")

style_label(ws7, 6, 1, "Decision name")
style_input(ws7, 6, 2, "Example: Take startup offer")

style_label(ws7, 8, 1, "OUTCOMES (up to 4)")
for c, h in enumerate(["Description", "Probability", "Payoff ($)", "Weighted"], 2):
    style_header(ws7, 9, c, h)

for i, (desc, prob, pay) in enumerate([
    ("Best case", 0.30, 300000),
    ("Good case", 0.30, 180000),
    ("Okay case", 0.25, 100000),
    ("Worst case", 0.15, 40000),
], start=10):
    style_label(ws7, i, 1, f"Outcome {i - 9}")
    style_input(ws7, i, 2, desc)
    style_input(ws7, i, 3, prob, "0%")
    style_input(ws7, i, 4, pay, "$#,##0")
    style_result(ws7, i, 5, formula=f"=C{i}*D{i}", fmt="$#,##0")

style_label(ws7, 14, 1, "Probabilities sum")
ws7.cell(row=14, column=3, value="=SUM(C10:C13)").number_format = "0%"
ws7.cell(row=14, column=3).border = THIN_BORDER
# Warn if not 100%
ws7.cell(row=14, column=4, value='=IF(ABS(C14-1)<0.001,"OK","WARNING: must = 100%")').font = Font(
    bold=True, color="C00000"
)

ws7.cell(row=16, column=1, value="RESULTS").font = Font(bold=True, size=13, color="2F5496")

# 1. EV
style_label(ws7, 18, 1, "1. EXPECTED VALUE")
style_result(ws7, 18, 2, formula="=SUM(E10:E13)", fmt="$#,##0")

# 2. Base rate
style_label(ws7, 20, 1, "2. BASE RATE CHECK")
style_label(ws7, 21, 1, "   Base rate for this type of decision")
style_input(ws7, 21, 2, 0.06, "0.0%")
ws7.cell(row=21, column=3, value="(e.g. 6% for startup success)").font = Font(size=9, color="999999")
style_label(ws7, 22, 1, "   Your estimate vs base rate")
style_result(ws7, 22, 2, formula='=IF(C10>B21,"You may be overconfident","Aligned with base rate")')

# 3. Sunk cost
style_label(ws7, 24, 1, "3. SUNK COST CHECK")
style_label(ws7, 25, 1, "   Already invested (ignore this!)")
style_input(ws7, 25, 2, 50000, "$#,##0")
ws7.cell(row=25, column=3, value="This should NOT affect your decision").font = Font(
    italic=True, size=9, color="C00000"
)

# 4. Bayesian prior
style_label(ws7, 27, 1, "4. BAYESIAN CONFIDENCE")
style_label(ws7, 28, 1, "   Your prior P(success)")
style_input(ws7, 28, 2, 0.50, "0%")
style_label(ws7, 29, 1, "   Strongest evidence for")
style_input(ws7, 29, 2, 0.80, "0%")
style_label(ws7, 30, 1, "   Same evidence if false")
style_input(ws7, 30, 2, 0.20, "0%")
style_label(ws7, 31, 1, "   Updated confidence")
style_result(ws7, 31, 2, formula="=(B29*B28)/(B29*B28+B30*(1-B28))", fmt="0.0%")

# 5. Survivorship
style_label(ws7, 33, 1, "5. SURVIVORSHIP CHECK")
style_label(ws7, 34, 1, "   Success stories you've seen")
style_input(ws7, 34, 2, 10)
style_label(ws7, 35, 1, "   Estimated total who tried")
style_input(ws7, 35, 2, 1000)
style_label(ws7, 36, 1, "   Real success rate")
style_result(ws7, 36, 2, formula="=B34/B35", fmt="0.0%")

# 6. Kelly
style_label(ws7, 38, 1, "6. KELLY SIZING")
style_label(ws7, 39, 1, "   Your bankroll / resources")
style_input(ws7, 39, 2, 100000, "$#,##0")
style_label(ws7, 40, 1, "   Win multiplier (gain/risk)")
style_input(ws7, 40, 2, 2.0, "0.0")
style_label(ws7, 41, 1, "   Full Kelly allocation")
style_result(ws7, 41, 2, formula="=MAX(0,(B31*B40-(1-B31))/B40)", fmt="0.0%")
style_label(ws7, 42, 1, "   Quarter Kelly (recommended)")
style_result(ws7, 42, 2, formula="=B41*0.25", fmt="0.0%")
style_label(ws7, 43, 1, "   Recommended commitment ($)")
style_result(ws7, 43, 2, formula="=B42*B39", fmt="$#,##0")

# Final verdict
ws7.cell(row=45, column=1, value="VERDICT").font = Font(bold=True, size=13, color="2F5496")
style_label(ws7, 46, 1, "Expected Value")
style_result(ws7, 46, 2, formula="=B18", fmt="$#,##0")
style_label(ws7, 47, 1, "Bayesian Confidence")
style_result(ws7, 47, 2, formula="=B31", fmt="0.0%")
style_label(ws7, 48, 1, "Recommended Bet Size")
style_result(ws7, 48, 2, formula="=B43", fmt="$#,##0")
style_label(ws7, 49, 1, "Go / No-Go")
style_result(ws7, 49, 2, formula='=IF(AND(B18>0,B31>0.5,B41>0),"GO","CAUTION - review inputs")')

# Move Dashboard to first position
wb.move_sheet("Dashboard", offset=-6)

# Save
wb.save("/Users/mayk/MathDecision/decision_math_formulas.xlsx")
print("Excel file created: decision_math_formulas.xlsx (7 sheets with charts)")
